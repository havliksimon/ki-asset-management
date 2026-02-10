"""
Comprehensive export utility for generating Excel reports with caching.

This module provides functionality to export all analyst data, performance metrics,
sector analysis, and charts to Excel format with multiple sheets.
"""

import os
import io
import logging
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Any
from functools import wraps

from ..extensions import db
from ..models import (
    Analysis, PerformanceCalculation, Company, User, Vote, 
    PortfolioPurchase, BenchmarkPrice, analysis_analysts
)
from ..utils.performance import PerformanceCalculator

logger = logging.getLogger(__name__)

# Cache configuration
EXPORT_CACHE_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'instance', 'exports')
CACHE_MAX_AGE_DAYS = 7


def ensure_export_dir():
    """Ensure export cache directory exists."""
    os.makedirs(EXPORT_CACHE_DIR, exist_ok=True)


def get_export_cache_path(export_type: str) -> str:
    """Get cache file path for export type."""
    ensure_export_dir()
    return os.path.join(EXPORT_CACHE_DIR, f'{export_type}_export.xlsx')


def get_export_metadata_path(export_type: str) -> str:
    """Get metadata file path for export."""
    ensure_export_dir()
    return os.path.join(EXPORT_CACHE_DIR, f'{export_type}_metadata.txt')


def get_cached_export_info(export_type: str) -> Optional[Dict[str, Any]]:
    """
    Get information about cached export if it exists.
    
    Returns:
        Dict with 'path', 'created_at', 'is_valid' or None if no cache
    """
    cache_path = get_export_cache_path(export_type)
    metadata_path = get_export_metadata_path(export_type)
    
    if not os.path.exists(cache_path) or not os.path.exists(metadata_path):
        return None
    
    try:
        with open(metadata_path, 'r') as f:
            created_at_str = f.read().strip()
            created_at = datetime.fromisoformat(created_at_str)
        
        age_days = (datetime.now() - created_at).days
        is_valid = age_days < CACHE_MAX_AGE_DAYS
        
        return {
            'path': cache_path,
            'created_at': created_at,
            'age_days': age_days,
            'is_valid': is_valid
        }
    except Exception as e:
        logger.warning(f"Error reading export metadata: {e}")
        return None


def save_export_cache(export_type: str, data: bytes):
    """Save export data to cache."""
    cache_path = get_export_cache_path(export_type)
    metadata_path = get_export_metadata_path(export_type)
    
    with open(cache_path, 'wb') as f:
        f.write(data)
    
    with open(metadata_path, 'w') as f:
        f.write(datetime.now().isoformat())


def generate_comprehensive_export(force_new: bool = False) -> tuple:
    """
    Generate comprehensive Excel export with all data.
    
    Returns:
        tuple: (bytes_data, is_from_cache, cache_info)
    """
    # Try to import required libraries first
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.chart import LineChart, Reference
        from openpyxl.drawing.image import Image as XLImage
    except ImportError as e:
        logger.error(f"pandas and openpyxl required for Excel export: {e}")
        raise ImportError("Excel export requires pandas and openpyxl. Please install: pip install pandas openpyxl")
    
    # Check cache first
    if not force_new:
        cache_info = get_cached_export_info('comprehensive')
        if cache_info and cache_info['is_valid']:
            with open(cache_info['path'], 'rb') as f:
                return f.read(), True, cache_info
    
    output = io.BytesIO()
    
    # Create Excel writer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        calculator = PerformanceCalculator()
        workbook = writer.book
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Sheet 1: Summary Dashboard
        summary_data = generate_summary_sheet(calculator)
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        _format_sheet(writer.sheets['Summary'], header_font, header_fill, header_alignment)
        
        # Sheet 2: All Analysts Performance
        analysts_data = generate_analysts_sheet(calculator)
        analysts_df = pd.DataFrame(analysts_data)
        analysts_df.to_excel(writer, sheet_name='Analyst Performance', index=False)
        _format_sheet(writer.sheets['Analyst Performance'], header_font, header_fill, header_alignment)
        
        # Sheet 3: All Analyses with Performance
        analyses_data = generate_analyses_sheet()
        analyses_df = pd.DataFrame(analyses_data)
        analyses_df.to_excel(writer, sheet_name='All Analyses', index=False)
        _format_sheet(writer.sheets['All Analyses'], header_font, header_fill, header_alignment)
        
        # Sheet 4: Board Portfolio
        board_data = generate_board_sheet()
        board_df = pd.DataFrame(board_data)
        board_df.to_excel(writer, sheet_name='Board Portfolio', index=False)
        _format_sheet(writer.sheets['Board Portfolio'], header_font, header_fill, header_alignment)
        
        # Sheet 5: Sector Analysis
        sector_data = generate_sector_sheet()
        sector_df = pd.DataFrame(sector_data)
        sector_df.to_excel(writer, sheet_name='Sector Analysis', index=False)
        _format_sheet(writer.sheets['Sector Analysis'], header_font, header_fill, header_alignment)
        
        # Sheet 6: Analyst Rankings
        rankings_data = generate_rankings_sheet(calculator)
        rankings_df = pd.DataFrame(rankings_data)
        rankings_df.to_excel(writer, sheet_name='Rankings', index=False)
        _format_sheet(writer.sheets['Rankings'], header_font, header_fill, header_alignment)
        
        # Sheet 7: Benchmark History
        benchmark_data = generate_benchmark_sheet()
        benchmark_df = pd.DataFrame(benchmark_data)
        benchmark_df.to_excel(writer, sheet_name='Benchmark History', index=False)
        _format_sheet(writer.sheets['Benchmark History'], header_font, header_fill, header_alignment)
        
        # Sheet 8: Detailed Returns (for chart recreation)
        returns_data = generate_detailed_returns_sheet()
        returns_df = pd.DataFrame(returns_data)
        returns_df.to_excel(writer, sheet_name='Detailed Returns', index=False)
        _format_sheet(writer.sheets['Detailed Returns'], header_font, header_fill, header_alignment)
        
        # Sheet 9-13: All 5 Overview Categories
        categories = [
            ('Purchased Only', 'purchased'),
            ('Board Approved', 'board_approved'),
            ('All Approved', 'all_approved'),
            ('Approved + Neutral', 'approved_neutral'),
            ('All Stocks', 'all')
        ]
        
        for cat_name, cat_filter in categories:
            cat_data = generate_overview_category_sheet(cat_filter)
            cat_df = pd.DataFrame(cat_data)
            sheet_name = f'Overview - {cat_name}'[:31]  # Excel sheet name limit
            cat_df.to_excel(writer, sheet_name=sheet_name, index=False)
            _format_sheet(writer.sheets[sheet_name], header_font, header_fill, header_alignment)
        
        # Sheet 14: Charts Data (for manual chart creation)
        charts_data = generate_charts_data_sheet()
        charts_df = pd.DataFrame(charts_data)
        charts_df.to_excel(writer, sheet_name='Charts Data', index=False)
        _format_sheet(writer.sheets['Charts Data'], header_font, header_fill, header_alignment)
    
    output.seek(0)
    data = output.read()
    
    # Save to cache
    save_export_cache('comprehensive', data)
    cache_info = get_cached_export_info('comprehensive')
    
    return data, False, cache_info


def _format_sheet(ws, header_font, header_fill, header_alignment):
    """Apply formatting to a worksheet."""
    # Format header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def generate_summary_sheet(calculator: PerformanceCalculator) -> List[Dict]:
    """Generate summary statistics."""
    from ..admin.routes import get_portfolio_performance
    
    portfolio = get_portfolio_performance(purchased_only=False)
    
    return [{
        'Metric': 'Export Generated',
        'Value': datetime.now().strftime('%Y-%m-%d %H:%M UTC'),
        'Notes': 'Comprehensive Analyst Report'
    }, {
        'Metric': 'Total Analysts',
        'Value': User.query.count(),
        'Notes': 'Active team members'
    }, {
        'Metric': 'Total Analyses',
        'Value': Analysis.query.count(),
        'Notes': 'All time'
    }, {
        'Metric': 'Approved Analyses',
        'Value': Analysis.query.filter_by(status='On Watchlist').count(),
        'Notes': 'On Watchlist'
    }, {
        'Metric': 'Board Positions',
        'Value': portfolio['num_positions'] if portfolio else 0,
        'Notes': 'Board approved'
    }, {
        'Metric': 'Portfolio Avg Return',
        'Value': f"{portfolio['total_return']}%" if portfolio else 'N/A',
        'Notes': 'Since inception'
    }, {
        'Metric': 'Portfolio Annualized',
        'Value': f"{portfolio['annualized_return']}%" if portfolio else 'N/A',
        'Notes': 'Annualized return'
    }, {
        'Metric': 'S&P 500 (1yr)',
        'Value': f"{portfolio['benchmark_spy']}%" if portfolio else 'N/A',
        'Notes': 'Benchmark'
    }, {
        'Metric': 'FTSE All-World (1yr)',
        'Value': f"{portfolio['benchmark_ftse']}%" if portfolio else 'N/A',
        'Notes': 'Benchmark'
    }, {
        'Metric': 'EEMS (1yr)',
        'Value': f"{portfolio['benchmark_eems']}%" if portfolio else 'N/A',
        'Notes': 'Benchmark'
    }]


def generate_analysts_sheet(calculator: PerformanceCalculator) -> List[Dict]:
    """Generate all analysts performance data."""
    data = []
    
    # Get all performance categories
    categories = [
        ('Approved Only', 'approved_only'),
        ('Neutral + Approved', 'neutral_approved'),
        ('All Stock Picks', 'all_stock')
    ]
    
    for cat_name, cat_filter in categories:
        all_perfs = calculator.get_all_analysts_performance(status_filter=cat_filter)
        
        for rank, perf in enumerate(all_perfs, 1):
            data.append({
                'Category': cat_name,
                'Rank': rank,
                'Analyst Name': perf['analyst_name'],
                'Analyst ID': perf['analyst_id'],
                'Number of Analyses': perf['num_analyses'],
                'Avg Return %': perf['avg_return'],
                'Median Return %': perf['median_return'],
                'Win Rate %': perf['win_rate'],
                'Best Return %': perf['best_return'],
                'Worst Return %': perf['worst_return']
            })
    
    return data


def generate_analyses_sheet() -> List[Dict]:
    """Generate all analyses with performance data."""
    data = []
    
    analyses = Analysis.query.all()
    
    for analysis in analyses:
        company = Company.query.get(analysis.company_id)
        
        # Get analysts and opponents
        analysts = [u.full_name or u.email for u in analysis.analysts_list]
        opponents = [u.full_name or u.email for u in analysis.opponents_list]
        
        # Get votes
        votes_yes = Vote.query.filter_by(analysis_id=analysis.id, vote=True).count()
        votes_no = Vote.query.filter_by(analysis_id=analysis.id, vote=False).count()
        
        # Get performance
        perf = PerformanceCalculation.query.filter_by(
            analysis_id=analysis.id
        ).order_by(PerformanceCalculation.calculation_date.desc()).first()
        
        # Check if purchased
        purchase = PortfolioPurchase.query.filter_by(analysis_id=analysis.id).first()
        
        data.append({
            'Analysis ID': analysis.id,
            'Company': company.name if company else 'Unknown',
            'Ticker': company.ticker_symbol if company else None,
            'Analysis Date': analysis.analysis_date,
            'Status': analysis.status,
            'Analysts': ', '.join(analysts),
            'Opponents': ', '.join(opponents),
            'Votes Yes': votes_yes,
            'Votes No': votes_no,
            'Board Approved': 'Yes' if votes_yes > votes_no else 'No',
            'Purchased': 'Yes' if purchase else 'No',
            'Purchase Date': purchase.purchase_date if purchase else None,
            'Price at Analysis': float(perf.price_at_analysis) if perf else None,
            'Current Price': float(perf.price_current) if perf else None,
            'Return %': float(perf.return_pct) if perf else None,
            'Last Updated': perf.calculation_date if perf else None
        })
    
    return data


def generate_board_sheet() -> List[Dict]:
    """Generate Board portfolio data."""
    data = []
    
    # Get all On Watchlist analyses
    analyses = Analysis.query.filter_by(status='On Watchlist').all()
    
    for analysis in analyses:
        company = Company.query.get(analysis.company_id)
        votes_yes = Vote.query.filter_by(analysis_id=analysis.id, vote=True).count()
        votes_no = Vote.query.filter_by(analysis_id=analysis.id, vote=False).count()
        purchase = PortfolioPurchase.query.filter_by(analysis_id=analysis.id).first()
        perf = PerformanceCalculation.query.filter_by(
            analysis_id=analysis.id
        ).order_by(PerformanceCalculation.calculation_date.desc()).first()
        
        data.append({
            'Company': company.name if company else 'Unknown',
            'Ticker': company.ticker_symbol if company else None,
            'Analysis Date': analysis.analysis_date,
            'Votes Yes': votes_yes,
            'Votes No': votes_no,
            'Status': 'Purchased' if purchase else 'Approved (Not Purchased)',
            'Purchase Date': purchase.purchase_date if purchase else None,
            'Return %': float(perf.return_pct) if perf else None,
            'Price at Analysis': float(perf.price_at_analysis) if perf else None,
            'Current Price': float(perf.price_current) if perf else None
        })
    
    return data


def generate_sector_sheet() -> List[Dict]:
    """Generate sector analysis data."""
    from ..utils.sector_helper import get_company_sector_async
    
    data = []
    sector_stats = {}
    
    analyses = Analysis.query.all()
    
    for analysis in analyses:
        company = Company.query.get(analysis.company_id)
        if not company or not company.ticker_symbol:
            continue
        
        sector = get_company_sector_async(company)
        if not sector:
            sector = 'Unknown'
        
        perf = PerformanceCalculation.query.filter_by(
            analysis_id=analysis.id
        ).order_by(PerformanceCalculation.calculation_date.desc()).first()
        
        if sector not in sector_stats:
            sector_stats[sector] = {
                'count': 0,
                'returns': [],
                'companies': []
            }
        
        sector_stats[sector]['count'] += 1
        sector_stats[sector]['companies'].append(company.name)
        
        if perf:
            sector_stats[sector]['returns'].append(float(perf.return_pct))
    
    for sector, stats in sector_stats.items():
        returns = stats['returns']
        if returns:
            avg_return = sum(returns) / len(returns)
            positive_count = sum(1 for r in returns if r > 0)
            positive_ratio = (positive_count / len(returns)) * 100
            
            if len(returns) > 1:
                mean = avg_return
                variance = sum((r - mean) ** 2 for r in returns) / len(returns)
                risk = variance ** 0.5
            else:
                risk = 0
            
            data.append({
                'Sector': sector,
                'Number of Stocks': stats['count'],
                'Avg Return %': round(avg_return, 2),
                'Positive Ratio %': round(positive_ratio, 2),
                'Risk (StdDev)': round(risk, 2),
                'Min Return %': round(min(returns), 2),
                'Max Return %': round(max(returns), 2),
                'Companies': ', '.join(stats['companies'][:10])  # First 10
            })
    
    # Sort by avg return
    data.sort(key=lambda x: x['Avg Return %'], reverse=True)
    
    return data


def generate_rankings_sheet(calculator: PerformanceCalculator) -> List[Dict]:
    """Generate analyst rankings."""
    data = []
    
    # Top by board approved
    board_counts = []
    for user in User.query.all():
        count = db.session.query(Analysis).join(
            analysis_analysts, Analysis.id == analysis_analysts.c.analysis_id
        ).filter(
            analysis_analysts.c.user_id == user.id,
            analysis_analysts.c.role == 'analyst',
            Analysis.status == 'On Watchlist'
        ).count()
        board_counts.append({
            'rank_type': 'Board Approved Count',
            'analyst': user.full_name or user.email.split('@')[0],
            'value': count,
            'metric': 'analyses'
        })
    
    board_counts.sort(key=lambda x: x['value'], reverse=True)
    for item in board_counts[:10]:
        data.append(item)
    
    # Top by total analyses (approved + neutral + refused)
    total_counts = []
    for user in User.query.all():
        count = db.session.query(Analysis).join(
            analysis_analysts, Analysis.id == analysis_analysts.c.analysis_id
        ).filter(
            analysis_analysts.c.user_id == user.id,
            analysis_analysts.c.role == 'analyst',
            Analysis.status.in_(['On Watchlist', 'Neutral', 'Refused'])
        ).count()
        total_counts.append({
            'rank_type': 'Total Analyses (All Stock Picks)',
            'analyst': user.full_name or user.email.split('@')[0],
            'value': count,
            'metric': 'analyses'
        })
    
    total_counts.sort(key=lambda x: x['value'], reverse=True)
    for item in total_counts[:10]:
        data.append(item)
    
    # Top by win rate
    all_perfs = calculator.get_all_analysts_performance()
    win_rates = []
    for perf in all_perfs:
        if perf['num_analyses'] >= 3 and perf['win_rate'] is not None:
            win_rates.append({
                'rank_type': 'Win Rate (3+ analyses)',
                'analyst': perf['analyst_name'],
                'value': perf['win_rate'],
                'metric': 'percent'
            })
    
    win_rates.sort(key=lambda x: x['value'], reverse=True)
    for item in win_rates[:10]:
        data.append(item)
    
    # Top by performance
    performances = []
    for perf in all_perfs:
        if perf['avg_return'] is not None:
            performances.append({
                'rank_type': 'Performance',
                'analyst': perf['analyst_name'],
                'value': perf['avg_return'],
                'metric': 'percent'
            })
    
    performances.sort(key=lambda x: x['value'], reverse=True)
    for item in performances[:10]:
        data.append(item)
    
    return data


def generate_benchmark_sheet() -> List[Dict]:
    """Generate benchmark price history."""
    data = []
    
    # Get last 2 years of benchmark data
    end_date = date.today()
    start_date = end_date - timedelta(days=730)
    
    benchmarks = BenchmarkPrice.query.filter(
        BenchmarkPrice.date >= start_date,
        BenchmarkPrice.date <= end_date
    ).order_by(BenchmarkPrice.date).all()
    
    for bp in benchmarks:
        data.append({
            'Date': bp.date,
            'Ticker': bp.ticker,
            'Close Price': float(bp.close_price),
            'Fetched At': bp.fetched_at
        })
    
    return data


def generate_detailed_returns_sheet() -> List[Dict]:
    """Generate detailed returns data for chart recreation."""
    data = []
    
    analyses = Analysis.query.all()
    
    for analysis in analyses:
        company = Company.query.get(analysis.company_id)
        perfs = PerformanceCalculation.query.filter_by(
            analysis_id=analysis.id
        ).order_by(PerformanceCalculation.calculation_date).all()
        
        for perf in perfs:
            data.append({
                'Analysis ID': analysis.id,
                'Company': company.name if company else 'Unknown',
                'Ticker': company.ticker_symbol if company else None,
                'Analysis Date': analysis.analysis_date,
                'Calculation Date': perf.calculation_date,
                'Price at Analysis': float(perf.price_at_analysis),
                'Price Current': float(perf.price_current),
                'Return %': float(perf.return_pct)
            })
    
    return data


def generate_overview_category_sheet(filter_type: str) -> List[Dict]:
    """Generate data for a specific overview category."""
    from ..admin.routes import get_portfolio_performance
    from ..analyst.routes import (
        get_analysis_ids_for_filter, get_portfolio_performance_for_analyses,
        get_portfolio_series_for_analyses
    )
    
    data = []
    
    # Get analysis IDs for this filter
    analysis_ids = get_analysis_ids_for_filter(filter_type)
    
    # Get performance
    perf = get_portfolio_performance_for_analyses(analysis_ids)
    
    # Get series data
    series_all = get_portfolio_series_for_analyses(analysis_ids, years=None)
    series_1y = get_portfolio_series_for_analyses(analysis_ids, years=1)
    
    # Summary row
    data.append({
        'Type': 'Summary',
        'Filter': filter_type.replace('_', ' ').title(),
        'Positions': perf.get('num_positions', 0),
        'Avg Return %': perf.get('total_return'),
        'Annualized %': perf.get('annualized_return'),
        'S&P 500 %': perf.get('benchmark_spy'),
        'FTSE %': perf.get('benchmark_ftse'),
        'EEMS %': perf.get('benchmark_eems'),
        'Inception': perf.get('start_date')
    })
    
    # Series data - Inception
    if series_all:
        for i, date_str in enumerate(series_all.get('dates', [])):
            data.append({
                'Type': 'Inception Series',
                'Date': date_str,
                'Portfolio %': series_all['portfolio_series'][i] if i < len(series_all['portfolio_series']) else None,
                'S&P 500 %': series_all['spy_series'][i] if i < len(series_all['spy_series']) else None,
                'FTSE %': series_all['vt_series'][i] if i < len(series_all['vt_series']) else None,
                'EEMS %': series_all['eems_series'][i] if i < len(series_all.get('eems_series', [])) else None
            })
    
    # Series data - 1 Year
    if series_1y:
        for i, date_str in enumerate(series_1y.get('dates', [])):
            data.append({
                'Type': '1Y Series',
                'Date': date_str,
                'Portfolio %': series_1y['portfolio_series'][i] if i < len(series_1y['portfolio_series']) else None,
                'S&P 500 %': series_1y['spy_series'][i] if i < len(series_1y['spy_series']) else None,
                'FTSE %': series_1y['vt_series'][i] if i < len(series_1y['vt_series']) else None,
                'EEMS %': series_1y['eems_series'][i] if i < len(series_1y.get('eems_series', [])) else None
            })
    
    return data


def generate_charts_data_sheet() -> List[Dict]:
    """Generate consolidated chart data for all categories."""
    from ..analyst.routes import (
        get_analysis_ids_for_filter, get_portfolio_series_for_analyses
    )
    
    data = []
    
    categories = [
        ('purchased', 'Purchased Only'),
        ('board_approved', 'Board Approved'),
        ('all_approved', 'All Approved'),
        ('approved_neutral', 'Approved + Neutral'),
        ('all', 'All Stocks')
    ]
    
    for filter_type, cat_name in categories:
        analysis_ids = get_analysis_ids_for_filter(filter_type)
        series_all = get_portfolio_series_for_analyses(analysis_ids, years=None)
        
        if series_all and series_all.get('dates'):
            for i, date_str in enumerate(series_all['dates']):
                data.append({
                    'Category': cat_name,
                    'Date': date_str,
                    'Portfolio': series_all['portfolio_series'][i] if i < len(series_all['portfolio_series']) else None,
                    'S&P 500': series_all['spy_series'][i] if i < len(series_all['spy_series']) else None,
                    'FTSE All-World': series_all['vt_series'][i] if i < len(series_all['vt_series']) else None,
                    'EEMS': series_all.get('eems_series', [None]*len(series_all['dates']))[i]
                })
    
    return data


def get_sector_cache_info() -> Optional[Dict[str, Any]]:
    """
    Get information about sector cache freshness.
    
    Returns:
        Dict with 'last_updated', 'age_days', 'is_fresh' or None
    """
    from ..models import CompanySectorCache
    
    latest_cache = CompanySectorCache.query.order_by(
        CompanySectorCache.fetched_at.desc()
    ).first()
    
    if not latest_cache:
        return None
    
    age_days = (datetime.utcnow() - latest_cache.fetched_at).days
    
    return {
        'last_updated': latest_cache.fetched_at,
        'age_days': age_days,
        'is_fresh': age_days <= 3
    }
